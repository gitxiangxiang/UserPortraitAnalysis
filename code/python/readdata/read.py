import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def read_data_from_excel(bookName, sheetName):
    """
    读取指定Excel的指定表单
    :param bookName:
    :param sheetName:
    :return:
    """
    # excel 路径
    #bookName = r"D:\文档\毕业设计数据\个人诉求.xlsx"
    # 表单名
    #sheetName = "第五人"

    # 打开Excel文件
    wb = openpyxl.load_workbook(bookName)
    # 选定表单
    sheet = wb[sheetName]
    # 获取行列数
    rowNum = sheet.max_row
    colNum = sheet.max_column
    if rowNum <= 1:
        return [[]]
    sheetContent = sheet[1:rowNum]
    if colNum > 10:
        # 最多只输出十条
        colNum = 10
    print("打印前十条投诉建议信息")
    for i in sheetContent[0:colNum]:
        for j in i:
            print(j.value, end=' ')
        print()
    return sheetContent


# def read_data_value_from_excel(table, sheet):
#     sheet_content = read_data_from_excel(table, sheet)
#     value = [j.value for i in sheet_content for j in i]
#     return value

if __name__ == '__main__':

    read_data_from_excel(r"..\..\..\数据\个人诉求.xlsx", '第五人')